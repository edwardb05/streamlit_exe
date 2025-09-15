import streamlit as st
import pandas as pd
from ortools.sat.python import cp_model
from rapidfuzz import process, fuzz
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import re
from dateutil.parser import parse
import time
import logging
import threading
import streamlit.components.v1 as components
from io import BytesIO
import pickle


# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

st.set_page_config(page_title="Exam Timetabling System", layout="wide")

# Core modules list
Core_modules = ["MECH70001 Nuclear Thermal Hydraulics",
                "MECH60004/MECH70042 Introduction to Nuclear Energy A/B",
                "MECH70002 Nuclear Reactor Physics",
                "MECH70008 Mechanical Transmissions Technology",
                "MECH70006 Metal Processing Technology",
                "MECH70021Aircraft Engine Technology",
                "MECH70003 Future Clean Transport Technology",
                "MECH60015/70030 PEN3/AME"]

# Fixed modules dictionary , name and date
Fixed_modules = {"BUSI60039 Business Strategy" :[1,1],
                 "BUSI60046 Project Management":[2,1],
                 "ME-ELEC70098 Optimisation":[3,0],
                 "MECH70001 Nuclear Thermal Hydraulics":[3,0],
                 "BUSI60040/BUSI60043 Corporate Finance Online/Finance & Financial Management":[3,1],
                 "MECH60004/MECH70042 Introduction to Nuclear Energy A/B":[4,0],
                 "ME-ELEC70022 Modelling and Control of Multi-body Mechanical Systems":[4,0],
                 "MATE97022 Nuclear Materials 1":[4,0],
                 "ME-MATE70029 Nuclear Fusion":[9,0],
                 "MECH70002 Nuclear Reactor Physics":[10,0],
                 "ME-ELEC70076 Sustainable Electrical Systems":[10,0],
                 "ME ELEC70066 Applied Advanced Optimisation":[10,0],
                 "MECH70020 Combustion, Safety and Fire Dynamics":[11,0],
                 "BIOE70016 Human Neuromechanical Control and Learning":[11,0],
                 "CENG60013 Nuclear Chemical Engineering":[11,0],
                 "MECH70008 Mechanical Transmissions Technology":[17,1],
                 "MECH70006 Metal Processing Technology":[17,1],
                 "MECH70021Aircraft Engine Technology":[17,1],
                 "MECH70003 Future Clean Transport Technology":[17,1],
                 "MECH60015/70030 PEN3/AME":[18,1]}

# Room dictionary with capacities and functions
rooms = {
    'CAGB 203': [["Computer", "SEQ"], 65],
    'CAGB 309': [["SEQ"], 54],
    'CAGB 649-652': [["SEQ"], 75],
    'CAGB 747-748': [["SEQ","AEA"], 36],
    'CAGB 749-752': [["SEQ"], 75],
    'CAGB 761': [["Computer", "SEQ","AEA"], 25],
    'CAGB 762': [["Computer", "SEQ","AEA"], 25],
    'CAGB 765': [["AEA","Computer"], 10],
    'CAGB 527': [["AEA"], 2],
    'NON ME N/A':[["SEQ","AEA"],1000], #For business and non Mech Eng modules
}

# No exam dates (weekends and last Friday morning)
no_exam_dates = [
    [5,0], [5,1], [6,0], [6,1],  # First weekend
    [12,0], [12,1], [13,0], [13,1],  # Second weekend
    [18,0], [19,0], [19,1], [20,0], [20,1]  # Last Friday morning and weekend
]

#Days it is preferable to not have an exam on but can if needed
no_exam_dates_soft = [
    [15,0],# Week 3 tuesday morning
    [16,0], #Week 3 Wednesday morning
]

def ordinal(n):
    # Returns ordinal string for an integer n, e.g. 1 -> 1st, 2 -> 2nd
    if 11 <= (n % 100) <= 13:
        return f"{n}th"
    else:
        return f"{n}{['th','st','nd','rd','th','th','th','th','th','th'][n % 10]}"

def validate_student_list(df):
    """Validate the student list Excel file format and content."""
    errors = []
    
    if len(df) < 3:
        errors.append("Student list must have at least 3 rows (header + students)")
        return errors
    
    if df.iloc[0, 0] != "CID" or df.iloc[0, 3] != "Additional Exam Arrangements AEA":
        errors.append(f"Student list must have 'CID' instead of {df.iloc[0, 0]} in column A and 'AEA' instead of {df.iloc[0, 3]}")
        return errors
    
    exam_columns = df.iloc[0, 9:].dropna()
    if len(exam_columns) == 0:
        errors.append("No exam columns found starting from column J")
        return errors
    
    student_rows = df.iloc[2:, :]
    for idx, row in student_rows.iterrows():
        cid = row[0]
        if pd.isna(cid):
            errors.append(f"Missing CID in row {idx + 3}")
            continue
        for col_idx, exam_name in enumerate(exam_columns, start=9):
            value = str(row[col_idx]).strip().lower()
            if value not in ['x', 'a', 'b', 'nan']:
                errors.append(f"Invalid exam indicator '{value}' for student {cid} in exam {exam_name}")

    return errors

def validate_module_list(df):
    """Validate the module list Excel file format and content."""
    errors = []

    if len(df) < 2:
        errors.append("Module list must have at least 2 rows")
        return errors
    
    required_cols = ['Banner Code (New CR)', 'Module Name', 'Module Leader (lecturer 1)']
    for col in required_cols:
        if col not in df.columns:
            errors.append(f"Missing required column: {col}")

    return errors

def validate_useful_dates(wb):
    """Validate the useful dates Excel file format and content."""
    errors = []
    if not wb:
        errors.append("Could not open useful dates file")
        return errors
    ws = wb.active
    found_bank_holidays = False
    row = 5
    while True:
        name = ws[f"F{row}"].value
        if name is None or "Term Dates" in str(name):
            break
        if "Bank Holiday" in str(name):
            found_bank_holidays = True
            break
        row += 1
    if not found_bank_holidays:
        errors.append("Could not find bank holidays section in useful dates file")
    found_summer_term = False
    row = 5
    while row < ws.max_row:
        cell_value = ws[f"F{row}"].value
        if cell_value and "Summer Term" in str(cell_value):
            found_summer_term = True
            break
        row += 1
    if not found_summer_term:
        errors.append("Could not find Summer Term section in useful dates file")
    return errors

def process_files():
    error = False
    """Process uploaded files and return processed data."""
    if not all([student_file, module_file, dates_file]):
        st.error("Please upload all required files")
        return None, None, None
    try:
        student_df = pd.read_excel(student_file, header=None)
        module_df = pd.read_excel(module_file, sheet_name=1, header=1)
        dates_wb = load_workbook(dates_file)
        student_errors = validate_student_list(student_df)
        if student_errors:
            st.error("Student list errors:\n" + "\n".join(student_errors))
            return None, None, None
        module_errors = validate_module_list(module_df)
        if module_errors:
            st.error("Module list errors:\n" + "\n".join(module_errors))
            return None, None, None
        dates_errors = validate_useful_dates(dates_wb)
        if dates_errors:
            st.error("Useful dates errors:\n" + "\n".join(dates_errors))
            return None, None, None
        
        #Read exams
        exams = student_df.iloc[0, 9:].dropna().tolist()
        student_rows = student_df.iloc[2:, :]
    
        #Form dictionary of each students exams
        student_exams = {}
        for _, row in student_rows.iterrows():
            cid = row[0]
            exams_taken = []
            for col_idx, exam_name in enumerate(exams, start=9):
                if str(row[col_idx]).strip().lower() in ['x', 'a', 'b']:
                    exams_taken.append(exam_name)
            student_exams[cid] = exams_taken
        for student in student_exams:
            for exam in student_exams[student]:
                if exam in Core_modules:
                    for other_exam in Fixed_modules:
                        if other_exam in student_exams[student]:
                            if exam != other_exam and Fixed_modules[exam][0] == Fixed_modules[other_exam][0]:
                                st.error(f"Core module {exam} conflicts with fixed module {other_exam} on the same day for student {student} so model will be infeasible")
                                error = True
        return student_df, module_df, dates_wb, error
    
    except Exception as e:
        st.error(f"Error processing files: {str(e)}")
        return None, None, None

def to_dict(obj):
    # Recursively convert defaultdicts to dicts
    if isinstance(obj, defaultdict):
        return dict((k, to_dict(v)) for k, v in obj.items())
    elif isinstance(obj, dict):
        return dict((k, to_dict(v)) for k, v in obj.items())
    elif isinstance(obj, list):
        return [to_dict(v) for v in obj]
    else:
        return obj

def create_timetable(students_df, leaders_df, wb,max_exams_2days, max_exams_5days):
    # Extract exam names from row 0, starting from column J (index 9)
    exams = students_df.iloc[0, 9:].dropna().tolist()
    # Get the range of rows containing student data (from row 3 onward)
    student_rows = students_df.iloc[2:, :]  # row index 3 and onward

    # Process bank holidays and create no_exam_dates
    ws = wb.active
    bank_holidays = []
    row = 5

    while True:
        name = ws[f"F{row}"].value
        date_cell = ws[f"G{row}"].value
        if name is None or "Term Dates" in str(name):
            break
        if isinstance(date_cell, datetime):
            bank_holidays.append((str(name).strip(), date_cell.date()))
        row += 1

    # Find Summer Term start date
    summer_start = None
    while row < ws.max_row:
        cell_value = ws[f"F{row}"].value
        if cell_value and "Summer Term" in str(cell_value):
            term_range = ws[f"F{row + 1}"].value
            if term_range:
                try:
                    start_part = term_range.split("to")[0].strip()
                    start_str = re.sub(r"^\w+\s+", "", start_part)
                    year_match = re.search(r"\b\d{4}\b", term_range)
                    if year_match:
                        start_str += f" {year_match.group(0)}"
                    else:
                        st.error("Year not found in date range.")
                        return None
                    summer_start = parse(start_str, dayfirst=True).date()
                except Exception as e:
                    st.error(f"Could not parse Summer Term start: {term_range}")
                    return None
            break
        row += 1
    if not summer_start:
        st.error("Summer Term start date not found")
        return None
    
    # Find first Monday
    first_monday = summer_start
    while first_monday.weekday() != 0:
        first_monday += timedelta(days=1)
    for name, bh_date in bank_holidays:
        delta = (bh_date - first_monday).days
        if 0 <= delta <= 20:
            no_exam_dates.append([delta, 0])
            no_exam_dates.append([delta, 1])

    #Form dictionary of student_exams
    student_exams = {}
    for _, row in student_rows.iterrows():
        cid = row[0]  # Column A = student CID
        exams_taken = []
        for col_idx, exam_name in enumerate(exams, start=9):
            if str(row[col_idx]).strip().lower() == 'x' or str(row[col_idx]).strip().lower() == 'a'  or str(row[col_idx]).strip().lower() == 'b' :  # Check for 'x' or 'a' or 'b' to indicate they take this course (case-insensitive)
                exams_taken.append(exam_name)
        student_exams[cid] = exams_taken
    student_rows = students_df.iloc[2:, :]  # row index 3 and onward
    
    #Get the list of days from useful dates
    days = []
    for i in range(21):
        date = first_monday + timedelta(days=i)
        day_str = date.strftime("%A ") + ordinal(date.day) + date.strftime(" %B")
        days.append(day_str)
    valid_aea_mask = (
        student_rows.iloc[:, 3].notna() &
        (student_rows.iloc[:, 3].astype(str).str.strip() != "#N/A")
    )

    AEA = student_rows.loc[valid_aea_mask, student_rows.columns[0]].tolist()
    
    standardized_names = exams

    leader_courses = defaultdict(list)
    exam_types = dict()

    for _, row in leaders_df.iterrows():
        leaders = []
        if pd.notna(row['Module Leader (lecturer 1)']):
            leaders.append(row['Module Leader (lecturer 1)'])
        if pd.notna(row['(UGO Internal) 2nd Exam Marker']):
            leaders.append(row['(UGO Internal) 2nd Exam Marker'])
        name = row['Module Name']
        code = row['Banner Code (New CR)']
        if pd.isna(code) or pd.isna(name) :
            continue
        if len(leaders) == 0 :
            continue
        combined_name = f"{code} {name}"
        best_match, score, _ = process.extractOne(
            combined_name, standardized_names, scorer=fuzz.token_sort_ratio
        )
        if score >= 70:
            exam_types[best_match] = row['(UGO Internal) Exam Style'] if pd.notna(row['(UGO Internal) Exam Style']) else None
            for leader in leaders:
                if best_match not in leader_courses[leader]:
                    leader_courses[leader].append(best_match)
    leader_courses = dict(leader_courses)


    for exam in exams:
        if exam not in exam_types:
            exam_types[exam] = "Standard"


    exam_counts = defaultdict(lambda: [0, 0])
    for cid, exams_taken in student_exams.items():
        if cid in AEA:
            for exam in exams_taken:
                exam_counts[exam][0] += 1
        else:
            for exam in exams_taken:
                exam_counts[exam][1] += 1

    exam_counts = dict(exam_counts)

    extra_time_students_25 = students_df[students_df.iloc[:, 3].astype(str).str.startswith(("15min/hour", "25% extra time"))].iloc[:, 0].tolist()
    extra_time_students_50 = students_df[students_df.iloc[:, 3].astype(str).str.startswith(("30min/hour", "50% extra time"))].iloc[:, 0].tolist()
    
    #####----- Start running the model----####
    model = cp_model.CpModel()
    slots = [0, 1]
    num_slots = len(slots)
    num_days = len(days)
    exam_day = {}
    exam_slot = {}
    for exam in exams:
        exam_day[exam] = model.NewIntVar(0, num_days - 1, f'{exam}_day')
        exam_slot[exam] = model.NewIntVar(0, num_slots - 1, f'{exam}_slot')
    exam_room = {}

    for exam in set().union(*student_exams.values()):
        for room in rooms:
            exam_room[(exam, room)] = model.NewBoolVar(f'{exam}_in_{room.replace(" ", "_")}')

#####----Adding constraints ------####
    # 0. Students can't have exams at the same tiem
    for student, exs in student_exams.items():
    #Loops through students
        for i in range(len(exs)):
            for j in range(i + 1, len(exs)):
                exam1 = exs[i]
                exam2 = exs[j]
                #Boolean variables for day and slot matches
                same_day = model.NewBoolVar(f'{exam1}_same_day{exam2}')
                same_slot = model.NewBoolVar(f'{exam1}_same_slot{exam2}')
                
                model.Add(exam_day[exam1] == exam_day[exam2]).OnlyEnforceIf(same_day)
                model.Add(exam_day[exam1] != exam_day[exam2]).OnlyEnforceIf(same_day.Not())

                model.Add(exam_slot[exam1] == exam_slot[exam2]).OnlyEnforceIf(same_slot)
                model.Add(exam_slot[exam1] != exam_slot[exam2]).OnlyEnforceIf(same_slot.Not())

                

                model.AddBoolOr([same_day.Not(), same_slot.Not()])


    # 1. Core modules can not have multiple exams on that day
    for student, exs in student_exams.items():
        core_mods = [exam for exam in exs if exam in Core_modules]
        other_mods = [exam for exam in exs if exam not in Core_modules]
        for exam in core_mods:
            for other in other_mods:
                model.Add(exam_day[exam] != exam_day[other])

    # 2. Fixed modules day and slot assignment
    for exam, (day_fixed, slot_fixed) in Fixed_modules.items():
        model.Add(exam_day[exam] == day_fixed)
        model.Add(exam_slot[exam] == slot_fixed)

    # 3. Forbidden exam day-slot assignments
    for exam in exams:
        for day, slot in no_exam_dates:
            model.AddForbiddenAssignments([exam_day[exam], exam_slot[exam]], [(day, slot)])
    # 4. Max 3 exams in any 2-day window per student
    for student, ex in student_exams.items():
        for d in range(num_days - 1):
            exams_in_2_days = []
            for exam in ex:
                is_on_d = model.NewBoolVar(f'{student}_{exam}_on_day_{d}')
                is_on_d1 = model.NewBoolVar(f'{student}_{exam}_on_day_{d+1}')
                is_on_either = model.NewBoolVar(f'{student}_{exam}_on_day_{d}_or_{d+1}')

                model.Add(exam_day[exam] == d).OnlyEnforceIf(is_on_d)
                model.Add(exam_day[exam] != d).OnlyEnforceIf(is_on_d.Not())

                model.Add(exam_day[exam] == d + 1).OnlyEnforceIf(is_on_d1)
                model.Add(exam_day[exam] != d + 1).OnlyEnforceIf(is_on_d1.Not())

                model.AddBoolOr([is_on_d, is_on_d1]).OnlyEnforceIf(is_on_either)
                model.AddBoolAnd([is_on_d.Not(), is_on_d1.Not()]).OnlyEnforceIf(is_on_either.Not())

                exams_in_2_days.append(is_on_either)

            model.Add(sum(exams_in_2_days) <= max_exams_2days)

    # 5. Max 4 exams in any 5-day sliding window per student
    for student, exs in student_exams.items():
        for start_day in range(num_days - 4):
            exams_in_window = []
            for exam in exs:
                in_window = model.NewBoolVar(f'{student}_{exam}_in_day_{start_day}_to_{start_day + 4}')

                model.AddLinearConstraint(exam_day[exam], start_day, start_day + 4).OnlyEnforceIf(in_window)

                before_window = model.NewBoolVar(f'{student}_{exam}_before_{start_day}')
                after_window = model.NewBoolVar(f'{student}_{exam}_after_{start_day + 4}')

                model.Add(exam_day[exam] < start_day).OnlyEnforceIf(before_window)
                model.Add(exam_day[exam] >= start_day).OnlyEnforceIf(before_window.Not())

                model.Add(exam_day[exam] > start_day + 4).OnlyEnforceIf(after_window)
                model.Add(exam_day[exam] <= start_day + 4).OnlyEnforceIf(after_window.Not())

                model.AddBoolOr([before_window, after_window]).OnlyEnforceIf(in_window.Not())

                exams_in_window.append(in_window)

            model.Add(sum(exams_in_window) <= max_exams_5days)

    # 6. At most 1 exam in week 3 (days 13 to 20) per module leader
    for leader, leader_exams in leader_courses.items():
        exams_in_week3 = []
        for exam in leader_exams:
            in_week3 = model.NewBoolVar(f'{exam}_in_week3')

            model.AddLinearConstraint(exam_day[exam], 13, 20).OnlyEnforceIf(in_week3)

            before_week3 = model.NewBoolVar(f'{exam}_before_week3')
            after_week3 = model.NewBoolVar(f'{exam}_after_week3')

            model.Add(exam_day[exam] < 13).OnlyEnforceIf(before_week3)
            model.Add(exam_day[exam] >= 13).OnlyEnforceIf(before_week3.Not())

            model.Add(exam_day[exam] > 20).OnlyEnforceIf(after_week3)
            model.Add(exam_day[exam] <= 20).OnlyEnforceIf(after_week3.Not())

            model.AddBoolOr([before_week3, after_week3]).OnlyEnforceIf(in_week3.Not())

            exams_in_week3.append(in_week3)

        model.Add(sum(exams_in_week3) <= 1)

    # 7. Extra time 50% students: max 1 exam per day
    for student in extra_time_students_50:
        for day in range(num_days):
            exams_on_day = []
            for exam in student_exams[student]:
                is_on_day = model.NewBoolVar(f'{student}_{exam}_on_day_{day}')
                model.Add(exam_day[exam] == day).OnlyEnforceIf(is_on_day)
                model.Add(exam_day[exam] != day).OnlyEnforceIf(is_on_day.Not())
                exams_on_day.append(is_on_day)
            model.Add(sum(exams_on_day) <= 1)

    #Soft constraint that extra time students with<= 25% should only have one a day
    extra_time_25_penalties= []
    for student in extra_time_students_25:
        for day in range(num_days):
            exams_on_day = []
            for exam in student_exams[student]:
                is_on_day = model.NewBoolVar(f'{student}_{exam}_on_day_{day}')
                model.Add(exam_day[exam] == day).OnlyEnforceIf(is_on_day)
                model.Add(exam_day[exam] != day).OnlyEnforceIf(is_on_day.Not())
                exams_on_day.append(is_on_day)
            num_exams = model.NewIntVar(0, len(exams_on_day), f'{student}_num_exams_day_{day}')
            model.Add(num_exams == sum(exams_on_day))
            has_multiple_exams = model.NewBoolVar(f'{student}_more_than_one_exam_day_{day}')
            model.Add(num_exams >= 2).OnlyEnforceIf(has_multiple_exams)
            model.Add(num_exams < 2).OnlyEnforceIf(has_multiple_exams.Not())
            penalty = model.NewIntVar(0, 5, f'{student}_penalty_day_{day}')
            model.Add(penalty == 5).OnlyEnforceIf(has_multiple_exams)
            model.Add(penalty == 0).OnlyEnforceIf(has_multiple_exams.Not())
            extra_time_25_penalties.append(penalty)

    #Soft constraint that course leaders modules should be spread out
    spread_penalties =[]
    for leader in leader_courses:
        mods = leader_courses[leader]
        for i in range(len(mods)):
            for j in range(i+1, len(mods)):
                m1 = mods[i]
                m2 = mods[j]
                diff = model.NewIntVar(-21, 21, f'{m1}_{m2}_diff')
                abs_diff = model.NewIntVar(0, 21, f'{m1}_{m2}_abs_diff')
                model.Add(diff == exam_day[m1] - exam_day[m2])
                model.AddAbsEquality(abs_diff, diff)
                close_penalty = model.NewIntVar(0, 5, f'{m1}_{m2}_penalty')
                is_gap_3 = model.NewBoolVar(f'{m1}_{m2}_gap3')
                is_gap_2 = model.NewBoolVar(f'{m1}_{m2}_gap2')
                is_gap_1 = model.NewBoolVar(f'{m1}_{m2}_gap1')
                is_gap_0 = model.NewBoolVar(f'{m1}_{m2}_gap0')
                model.Add(abs_diff == 3).OnlyEnforceIf(is_gap_3)
                model.Add(abs_diff != 3).OnlyEnforceIf(is_gap_3.Not())
                model.Add(abs_diff == 2).OnlyEnforceIf(is_gap_2)
                model.Add(abs_diff != 2).OnlyEnforceIf(is_gap_2.Not())
                model.Add(abs_diff == 1).OnlyEnforceIf(is_gap_1)
                model.Add(abs_diff != 1).OnlyEnforceIf(is_gap_1.Not())
                model.Add(abs_diff == 0).OnlyEnforceIf(is_gap_0)
                model.Add(abs_diff != 0).OnlyEnforceIf(is_gap_0.Not())
                model.Add(close_penalty == 1).OnlyEnforceIf(is_gap_3)
                model.Add(close_penalty == 3).OnlyEnforceIf(is_gap_2)
                model.Add(close_penalty == 4).OnlyEnforceIf(is_gap_1)
                model.Add(close_penalty == 5).OnlyEnforceIf(is_gap_0)
                model.Add(close_penalty == 0).OnlyEnforceIf(
                    is_gap_3.Not(), is_gap_2.Not(), is_gap_1.Not(), is_gap_0.Not()
                )
                spread_penalties.append(close_penalty)

    #Soft constraint to ensure no exams on some days
    soft_day_penalties = []
    for exam in exams:
        for day, slot in no_exam_dates_soft:
            is_on_soft_day = model.NewBoolVar(f'{exam}_on_soft_day_{day}_{slot}')
            day_match = model.NewBoolVar(f'{exam}_day_eq_{day}')
            slot_match = model.NewBoolVar(f'{exam}_slot_eq_{slot}')
            model.Add(exam_day[exam] == day).OnlyEnforceIf(day_match)
            model.Add(exam_day[exam] != day).OnlyEnforceIf(day_match.Not())
            model.Add(exam_slot[exam] == slot).OnlyEnforceIf(slot_match)
            model.Add(exam_slot[exam] != slot).OnlyEnforceIf(slot_match.Not())
            model.AddBoolAnd([day_match, slot_match]).OnlyEnforceIf(is_on_soft_day)
            model.AddBoolOr([day_match.Not(), slot_match.Not()]).OnlyEnforceIf(is_on_soft_day.Not())
            penalty = model.NewIntVar(0, 5, f'{exam}_penalty_soft_day_{day}_{slot}')
            model.Add(penalty == 5).OnlyEnforceIf(is_on_soft_day)
            model.Add(penalty == 0).OnlyEnforceIf(is_on_soft_day.Not())
            soft_day_penalties.append(penalty)

    #Minimize the amount of exams per slot 
    soft_slot_penalties = []

    for day in range(15):  #1 First two weeks only
        for slot in slots:  
            exams_in_slot = []
                
                    # 2 Make a list of all exams in a slot
            for exam in exams:
                is_scheduled_day = model.NewBoolVar(f'{exam}_is_on_day{day}')
                is_scheduled_slot = model.NewBoolVar(f'{exam}_is_on_slot{slot}')

                model.Add(exam_day[exam] == day).OnlyEnforceIf(is_scheduled_day)
                model.Add(exam_day[exam] != day).OnlyEnforceIf(is_scheduled_day.Not())

                model.Add(exam_slot[exam] == slot).OnlyEnforceIf(is_scheduled_slot)
                model.Add(exam_slot[exam] != slot).OnlyEnforceIf(is_scheduled_slot.Not())
                
                is_scheduled_here = model.NewBoolVar(f'{exam}_on_day{day}_slot{slot}')
                model.AddBoolAnd([is_scheduled_day, is_scheduled_slot]).OnlyEnforceIf(is_scheduled_here)
                model.AddBoolOr([is_scheduled_day.Not(), is_scheduled_slot.Not()]).OnlyEnforceIf(is_scheduled_here.Not())
                exams_in_slot.append(is_scheduled_here)

            # 3 Count number of exams scheduled in this (day, slot)
            num_exams_here = model.NewIntVar(0, len(exams), f'count_day{day}_slot{slot}')
            model.Add(num_exams_here == sum(exams_in_slot))

            # 4 Calculate penalties
            is_three = model.NewBoolVar(f'is_three_day{day}_slot{slot}')
            is_four_or_more = model.NewBoolVar(f'is_four_plus_day{day}_slot{slot}')

            model.Add(num_exams_here == 3).OnlyEnforceIf(is_three)
            model.Add(num_exams_here != 3).OnlyEnforceIf(is_three.Not())

            model.Add(num_exams_here >= 4).OnlyEnforceIf(is_four_or_more)
            model.Add(num_exams_here < 4).OnlyEnforceIf(is_four_or_more.Not())

            #5 Apply penalties
            penalty_three = model.NewIntVar(0, 5, f'penalty_three_day{day}_slot{slot}')
            penalty_four = model.NewIntVar(0, 10, f'penalty_four_day{day}_slot{slot}')

            model.Add(penalty_three == 5).OnlyEnforceIf(is_three)
            model.Add(penalty_three == 0).OnlyEnforceIf(is_three.Not())

            model.Add(penalty_four == 10).OnlyEnforceIf(is_four_or_more)
            model.Add(penalty_four == 0).OnlyEnforceIf(is_four_or_more.Not())

            soft_slot_penalties.append(penalty_three)
            soft_slot_penalties.append(penalty_four)

   ####- room constraints - ####
    # Ensure each non ME exam is assigned room N/A and ME is not assingned this
    for exam in exams:
        if exam in Fixed_modules and exam not in Core_modules:
            model.Add(exam_room[(exam, 'NON ME N/A')] ==1)  # Assign to N/A room if fixed module
        else:
            model.Add(exam_room[(exam, 'NON ME N/A')] == 0)  # Do not assign to N/A room if not fixed module


    #Must have sufficient room for each exam 
    for exam in exams:

        AEA_capacity = sum(
            rooms[room][1] * exam_room[(exam, room)]
            for room in rooms if "AEA" in rooms[room][0]
        )
        SEQ_capacity = sum(
            rooms[room][1] * exam_room[(exam, room)]
            for room in rooms if "SEQ" in rooms[room][0]
        )
        AEA_students = exam_counts[exam][0]
        SEQ_students = exam_counts[exam][1]
        model.Add(AEA_capacity >= AEA_students)
        model.Add(SEQ_capacity >= SEQ_students)

    #Ensure only one day and slot assigned to each room
    for d in range(num_days):
        for s in range(num_slots):
            for room in rooms:
                if room == 'NON ME N/A':
                    continue  # Skip N/A room for this constraint 
                else:
                    exams_in_room_time = []
                    for exam in exams:
                        exam_at_day = model.NewBoolVar(f'{exam}_on_day_{d}')
                        model.Add(exam_day[exam] == d).OnlyEnforceIf(exam_at_day)
                        model.Add(exam_day[exam] != d).OnlyEnforceIf(exam_at_day.Not())

                        exam_at_slot = model.NewBoolVar(f'{exam}_on_slot_{s}')
                        model.Add(exam_slot[exam] == s).OnlyEnforceIf(exam_at_slot)
                        model.Add(exam_slot[exam] != s).OnlyEnforceIf(exam_at_slot.Not())

                        exam_at_time = model.NewBoolVar(f'{exam}_on_{d}_{s}')
                        model.AddBoolAnd([exam_at_day, exam_at_slot]).OnlyEnforceIf(exam_at_time)
                        model.AddBoolOr([exam_at_day.Not(), exam_at_slot.Not()]).OnlyEnforceIf(exam_at_time.Not())

                        assigned_and_scheduled = model.NewBoolVar(f'{exam}_in_{room}_at_{d}_{s}')
                        model.AddBoolAnd([exam_room[(exam, room)], exam_at_time]).OnlyEnforceIf(assigned_and_scheduled)
                        model.AddBoolOr([exam_room[(exam, room)].Not(), exam_at_time.Not()]).OnlyEnforceIf(assigned_and_scheduled.Not())

                        exams_in_room_time.append(assigned_and_scheduled)
                    model.AddAtMostOne(exams_in_room_time)

    #Ensure non computer rooms not used for computer exams
    for exam in exams:
        if exam_types[exam] == "PC":
            for room in rooms:
                uses = rooms[room][0]
                if "Computer" not in uses:
                    model.Add(exam_room[(exam, room)] == 0)

    # Minimize amount of rooms used
    room_surplus = []
    for exam in exams:
        model.Add(sum(exam_room[(exam, room)] for room in rooms) >= 1)
        rooms_len = model.NewIntVar(0, 9, f'rooms for {exam}')

        model.Add(rooms_len == sum(exam_room[(exam, room)]for room in rooms))
        rooms_penalty = model.NewIntVar(0, 15, f'{exam}_room_surplus_penalty')

        is_room_length_greater_6 = model.NewBoolVar(f'{exam}_has_six_or_more_rooms')
        is_room_length_5 = model.NewBoolVar(f'{exam}_has_five_rooms')
        is_room_length_4 = model.NewBoolVar(f'{exam}_has_four_rooms')
        is_room_length_3 = model.NewBoolVar(f'{exam}_has_three_rooms')

        model.Add(rooms_len >= 6).OnlyEnforceIf(is_room_length_greater_6)
        model.Add(rooms_len <= 5).OnlyEnforceIf(is_room_length_greater_6.Not())
        model.Add(rooms_len == 5).OnlyEnforceIf(is_room_length_5)
        model.Add(rooms_len != 5).OnlyEnforceIf(is_room_length_5.Not())
        model.Add(rooms_len == 4).OnlyEnforceIf(is_room_length_4)
        model.Add(rooms_len != 4).OnlyEnforceIf(is_room_length_4.Not())
        model.Add(rooms_len == 3).OnlyEnforceIf(is_room_length_3)
        model.Add(rooms_len != 3).OnlyEnforceIf(is_room_length_3.Not())

        model.add(rooms_penalty == 15).OnlyEnforceIf(is_room_length_greater_6)
        model.Add(rooms_penalty == 9).OnlyEnforceIf(is_room_length_5)
        model.Add(rooms_penalty == 6).OnlyEnforceIf(is_room_length_4)
        model.Add(rooms_penalty == 4).OnlyEnforceIf(is_room_length_3)
        model.Add(rooms_penalty == 0).OnlyEnforceIf(
                    is_room_length_3.Not(), is_room_length_4.Not(), is_room_length_5.Not(), is_room_length_greater_6.Not(),
                )
        room_surplus.append(rooms_penalty)
    

    #Penalise using pc rooms for non pc exams

    non_pc_exam_penalty = []

    #1 Find computer rooms
    computer_rooms = [room for room in rooms if "Computer" in rooms[room][0]]

    #2 Loop through exams
    for exam in exams:
            #3 if not a PC exams
        if exam_types[exam] != "PC":
                #4 Check each computer room
            for room in computer_rooms:
                #5 Create a Boolean 
                penalty_var = model.NewBoolVar(f"non_pc_exam_in_pc_room_{exam}_{room}")
                
                #6 Assign penalty 
                model.Add(exam_room[(exam, room)] == 1).OnlyEnforceIf(penalty_var)
                model.Add(exam_room[(exam, room)] != 1).OnlyEnforceIf(penalty_var.Not())

                #7 Add penality
                non_pc_exam_penalty.append(5 * penalty_var)
            
    model.Minimize(sum(spread_penalties) + sum(soft_day_penalties)*soft_day_penalty+   sum(extra_time_25_penalties)*extra_time_penalty+sum(room_surplus)+ sum(soft_slot_penalties)*soft_slot_penalty+ sum(non_pc_exam_penalty))
   
    #### ----- Solve the model ----- ###
    solver = cp_model.CpSolver()
    # set max time
    solver.parameters.max_time_in_seconds = 120 
    status = solver.Solve(model)
    if status == cp_model.FEASIBLE or status == cp_model.OPTIMAL:
        exams_timetabled = {}
        for exam in exams:
            d = solver.Value(exam_day[exam])
            s = solver.Value(exam_slot[exam])
            assigned_rooms = [room for room in rooms if solver.Value(exam_room[(exam, room)]) == 1]
            try:
                leader = [name for name, exams in leader_courses.items() if exam in exams][0]
            except IndexError:
                leader = "unknown"
            exams_timetabled[exam] = (d, s, assigned_rooms)

        # Save data for next page to pickle - unpacked when generation finished and in main scope
        data_to_save ={
            "days": days,
            "slots": [0, 1],
            "exams": exams,
            "AEA": AEA,
            "leader_courses": leader_courses,
            "extra_time_students_25": extra_time_students_25,
            "extra_time_students_50": extra_time_students_50,
            "student_exams": student_exams,
            "exam_counts": exam_counts,
            "Fixed_modules": Fixed_modules,
            "Core_modules": Core_modules,
            "rooms": rooms,
            "exam_types": exam_types,
        }
        pickle_buffer = BytesIO()
        pickle.dump(data_to_save, pickle_buffer)
        pickle_buffer.seek(0)
        total_penalty = sum(solver.Value(v) for v in spread_penalties + soft_day_penalties + room_surplus +extra_time_25_penalties)
        return exams_timetabled, days, exam_counts, exam_types,total_penalty, pickle_buffer
    
    elif status == cp_model.INFEASIBLE:
        # print infeasible boolean variables index
        st.error('Infeasible model. Exam schedule could not be created.')
    else:
        st.error("No solution found.")



def generate_excel(exams_timetabled, days, exam_counts, exam_types):
    # ------------ BUILD rows and row_meta ------------
    data = {}
    for exam, (d, s, room, *_) in exams_timetabled.items():
        day = days[d]
        slot = s
        data.setdefault(day, {}).setdefault(slot, []).append((exam, room))

    rows = []
    row_meta = []

    for d_idx, day_name in enumerate(days):
        for s_idx, slot_name in enumerate(['Morning', 'Afternoon']):
            exams_list = data.get(day_name, {}).get(s_idx, [])
            if not exams_list:
                rows.append([day_name, slot_name, '', '', '', ''])
                row_meta.append((d_idx, s_idx))
            else:
                for exam_name, room in exams_list:
                    room_str = ', '.join(room)
                    total_students = f'AEA {exam_counts[exam_name][0]}, Non-AEA {exam_counts[exam_name][1]}'
                    type_str = " (Computer)" if exam_types[exam_name] == "PC" else " (Standard)"
                    rows.append([day_name, slot_name, exam_name, total_students, room_str, type_str])
                    row_meta.append((d_idx, s_idx))
                rows.append([day_name, slot_name, '', '', '', ''])
                row_meta.append((d_idx, s_idx))

    df = pd.DataFrame(rows, columns=['Date', 'Time', 'Exam', 'Total No of Students', 'Room', 'Type'])

    # ------------ CREATE workbook and append rows ------------
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # ------------ FUNCTION to merge vertical cells ------------
    def merge_vertical(col, key_fn):
        start = 2
        last_key = key_fn(start)
        for r in range(3, ws.max_row + 2):
            key = key_fn(r) if r <= ws.max_row else None
            if key != last_key:
                if r - start > 1:
                    ws.merge_cells(start_row=start, start_column=col, end_row=r-1, end_column=col)
                start = r
                last_key = key

    # Merge Time cells: consecutive identical (Date, Time) pairs
    merge_vertical(2, lambda r: (ws.cell(r,1).value, ws.cell(r,2).value))
    # Merge Date cells: all rows for the same day
    merge_vertical(1, lambda r: ws.cell(r, 1).value)

    # ------------ DEFINE fills ------------
    yellow = PatternFill('solid', fgColor='FFFF54')  # Fixed modules
    red = PatternFill('solid', fgColor='EA3323')     # Core modules
    blue = PatternFill('solid', fgColor='E0EAF6')    # Alternating row color
    green = PatternFill('solid', fgColor='CBE9B8')   # Alternating row color

    # ------------ APPLY alternating row fills BY DAY ------------
    for excel_row, (d_idx, s_idx) in enumerate(row_meta, start=2):
        fill = blue if d_idx % 2 == 0 else green
        for col in range(1, 7):
            ws.cell(row=excel_row, column=col).fill = fill

    # ------------ APPLY fixed/core exam coloring ------------
    for r in range(2, ws.max_row + 1):
        exam_name = ws.cell(r, 3).value
        fill = None
        if exam_name:
            if any(exam_name.startswith(fm) for fm in Fixed_modules):
                fill = yellow
            if any(exam_name.startswith(cm) for cm in Core_modules):
                fill = red
        if fill:
            for c in (3, 4, 5, 6):
                ws.cell(r, c).fill = fill

    # ------------ CENTER text for Date and Time columns ------------
    for row in range(2, ws.max_row + 1):
        for col in [1, 2]:
            ws.cell(row=row, column=col).alignment = Alignment(vertical='center')

    # ------------ AUTO-WIDTH columns ------------
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # ------------ SAVE workbook to BytesIO(temporary storage) ------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

#Rotating filling animation
def animation_html():
    return """
    <div class=\"wrapper\">
      <div class=\"container\" id=\"container\">
        <div class=\"flange-top\">
          <div class=\"flange-fill\" id=\"flangeTopFill\"></div>
        </div>
        <div class=\"flange-bottom\">
          <div class=\"flange-fill\" id=\"flangeBottomFill\"></div>
        </div>
        <div class=\"i-body\">
          <div class=\"fill\" id=\"bodyFill\"></div>
        </div>
      </div>
    </div>

<style>
  .wrapper {
    width: 267px;
    height: 267px;
    background-color: #00ff7f;
    border-radius: 50%;
    display: flex;
    justify-content: center;
    align-items: center;
    margin: 40px auto;
    overflow: hidden;
  }

  .container {
    width: 267px;
    height: 267px;
    position: relative;
    transform-origin: center;
    transition: transform 1s ease-in-out;
  }

  .flange-top,
  .flange-bottom {
    width: 117px;
    height: 30px;
    background: white;
    position: absolute;
    left: 75px;
    overflow: hidden;
  }

  .flange-top {
    top: 40px;
  }

  .flange-bottom {
    bottom: 40px;
  }

  .flange-fill {
    position: absolute;
    bottom: 0;
    left: 0;
    width: 100%;
    height: 0%;
    background: #0000cd;
    transition: height 1s ease-in-out;
  }

  .i-body {
    position: absolute;
    top: 70px;
    height: 127px;
    width: 31px;
    left: 118px;
    background: white;
    overflow: hidden;
    box-sizing: border-box;
  }

  .fill {
    position: absolute;
    bottom: 0;
    width: 100%;
    height: 0%;
    background: #0000cd;
    transition: height 1s ease-in-out;
  }
</style>


<script>
  let angle = 0;
  const delay = (ms) => new Promise(resolve => setTimeout(resolve, ms));

  const container = document.getElementById("container");
  const bodyFill = document.getElementById("bodyFill");
  const flangeTopFill = document.getElementById("flangeTopFill");
  const flangeBottomFill = document.getElementById("flangeBottomFill");

  function setFillDirection(element, fromTop) {
    if (fromTop) {
      element.style.top = '0';
      element.style.bottom = '';
    } else {
      element.style.bottom = '0';
      element.style.top = '';
    }
  }

  async function fillSequence1(fromTop) {
    setFillDirection(flangeBottomFill, fromTop);
    flangeBottomFill.style.height = "100%";
    await delay(1000);

    setFillDirection(bodyFill, fromTop);
    bodyFill.style.height = "100%";
    await delay(1000);

    setFillDirection(flangeTopFill, fromTop);
    flangeTopFill.style.height = "100%";
    await delay(1000);
  }

  async function emptySequence1(fromTop) {
    setFillDirection(flangeTopFill, fromTop);
    flangeTopFill.style.height = "0%";
    await delay(1000);

    setFillDirection(bodyFill, fromTop);
    bodyFill.style.height = "0%";
    await delay(1000);

    setFillDirection(flangeBottomFill, fromTop);
    flangeBottomFill.style.height = "0%";
    await delay(1000);
  }
   async function fillSequence2(fromTop) {
    setFillDirection(flangeTopFill, fromTop);
    flangeTopFill.style.height = "100%";
    await delay(1000);



    setFillDirection(bodyFill, fromTop);
    bodyFill.style.height = "100%";
    await delay(1000);

    setFillDirection(flangeBottomFill, fromTop);
    flangeBottomFill.style.height = "100%";
    await delay(1000);
  }

  async function emptySequence2(fromTop) {
    setFillDirection(flangeBottomFill, fromTop);
    flangeBottomFill.style.height = "0%";
    await delay(1000);


    setFillDirection(bodyFill, fromTop);
    bodyFill.style.height = "0%";
    await delay(1000);

    setFillDirection(flangeTopFill, fromTop);
    flangeTopFill.style.height = "0%";
    await delay(1000);
  }

  async function animateCycle() {
    while (true) {
      // Fill bottom-to-top visually (normal orientation)
      await fillSequence1(false);
      await delay(500);

      angle += 180;
      container.style.transform = `rotate(${angle}deg)`;
      await delay(1000);

      // Empty bottom-to-top visually (but now rotated, so DOM-top is visual-bottom)
      await emptySequence2(true);
      await delay(500);

      // Fill bottom-to-top visually (still rotated, so fill from top in DOM)
      await fillSequence2(true);
      await delay(500);

      angle += 180;
      container.style.transform = `rotate(${angle}deg)`;
      await delay(1000);

      // Empty bottom-to-top visually (now upright, so DOM-bottom is visual-bottom)
      await emptySequence1(false);
      await delay(500);
    }
  }

  // Start animation after DOM ready
  setTimeout(() => {
    animateCycle();
  }, 100);
</script>
    """

# --- MAIN STREAMLIT UI LOGIC ---
st.title("Exam Timetabling System")

# File upload section
st.header("Upload Required Files")
col1, col2, col3 = st.columns(3)

with col1:
    student_file = st.file_uploader("Upload Student List", type=['xlsx'])
with col2:
    module_file = st.file_uploader("Upload Module List", type=['xlsx'])
with col3:
    dates_file = st.file_uploader("Upload Useful Dates", type=['xlsx'])

# Parameters section
st.header("Timetabling Parameters")
st.markdown(""" Adjust the parameters below to customize the exam scheduling process. These parameters will influence how the exams are distributed across the available days and slots.
            The sliders on the right represent the weighting of the soft constraints, which can be adjusted to prioritize certain aspects of the timetable generation process.""")
col1, col2 = st.columns(2)

with col1:
    num_days = st.number_input("Number of Days for Exam Period", min_value=1, max_value=30, value=21) -1 # Subtract 1 to match the 0-indexed days in the code
    max_exams_2days = st.number_input("Maximum Exams in 2-Day Window", min_value=1, max_value=5, value=3)
    max_exams_5days = st.number_input("Maximum Exams in 5-Day Window", min_value=1, max_value=10, value=4)

with col2:
    room_penalty = st.slider("Having non PC exams in computer room penalty weight", min_value=0, max_value=10, value=5)/5 #divide by 5 to normalize it 
    extra_time_penalty = st.slider(r"25% Extra Time Students having more than one exam a day Penalty Weight", min_value=0, max_value=10, value=5)/5
    soft_day_penalty = st.slider("Soft constraint for no exams on certain days (Week 3 Tuesday and Wednesdnay Morning) Penalty Weight", min_value=0, max_value=10, value=5)/5

# Add a generate button
if st.button("Generate Timetable"):
    students_df, leaders_df, wb, error = process_files()
    if not all([student_file, module_file, dates_file]):
        st.error("Please upload all required files first.")
    elif error is True:
        st.error("Please ensure files are fixed before trying again.")
    else:
        try:
            animation_placeholder = st.empty()
            result_container = st.empty()
            processing_done = False
            error_msg = None
            output = None
            def generate():
                global processing_done, error_msg, students_df, leaders_df, penalties, output, pickle_buffer
                try:
                    timetable, days, exam_counts, exam_types, penalties, pickle_buffer = create_timetable(
                        students_df, leaders_df, wb, max_exams_2days, max_exams_5days,
                    )
                    output = generate_excel(timetable, days, exam_counts, exam_types)


                except Exception as e:
                    error_msg = str(e)
                finally:
                    processing_done = True
            thread = threading.Thread(target=generate)
            thread.start()

            while not processing_done:
                with animation_placeholder:
                    components.html(animation_html(), height=350)
                time.sleep(2.1)

            animation_placeholder.empty()

            if error_msg:
                st.error(f"An error occurred: {error_msg}")
                logger.error(f"Error generating timetable: {error_msg}", exc_info=True)
            else:
                pickle_buffer.seek(0)
                data = pickle.load(pickle_buffer)
                #Unpack the data from the temporary pickle file and write it to session state to carry across pages
                st.session_state["exam_data"] = data
                st.success("✅ Timetable generated successfully!")
                st.write(f"Total Penalty: {penalties}")
                st.download_button(
                    label="Download Timetable",
                    data=output,
                    file_name="exam_schedule.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.header("Generated Timetable")
                df = pd.read_excel(output)
                st.dataframe(df)
        except Exception as e:
            st.error(f"Unexpected error: {str(e)}")

